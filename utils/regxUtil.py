import os
import re
import pandas as pd
from common.get_data import Getdata
from utils.operation_excel import OperationExcel
from jsonpath_rw import jsonpath, parse
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
import xlwt

# s='www.lemfix.com'
# res = re.match('(w)(ww)',s)
# #group()==group(0)拿到匹配的全字符;分组根据你正则表达式里面的括号去分组
# print(res.group(0))
# print(res.group(1))
# print(res.group(2))

# a = 'hellolemonfixlemon'
# res = re.findall('(le)(mon)',a)#返回列表 在字符串里面找 匹配的内容 存在列表里面
# #如果有分组 就是义元组的形式表现出来 列表嵌套元组
# print(res)
#
# b = '{"moblie":"${moblie}","pwd":"${pwd}"}'
# res = re.search('\$\{(.*?)\}',b)
# print(res)
# print(res.group(0))
# print(res.group(1))

class DoRegx:
    # 利用反射向Getdata的变量取值完成变量替换
    def do_regx1(self,s):
        while re.search('\$\{(.*?)\}',s):
            key = re.search('\$\{(.*?)\}',s).group(0)
            value = re.search('\$\{(.*?)\}', s).group(1)
            s = s.replace(key,str(getattr(Getdata,value)))
        print(s)
        return s


    def do_regx(self,s):
        while re.search('\$\{(.*?)\}',s):
            key = re.search('\$\{(.*?)\}',s).group(0)
            value = re.search('\$\{(.*?)\}', s).group(1)
            s = s.replace(key,str(getattr(Getdata,value)))
            if value=='name':
                setattr(Getdata, value,DoRegx().incr14Str(str(getattr(Getdata, value))))
                DoRegx().updata_init(Getdata.case_path,'init',str(getattr(Getdata,value)))
        print(s)
        return s

    # 利用反射给Getdata的变量赋值
    def get_parmvalue(self,case_no,sheet_name):
        casepath=r'C:\Users\wangchao\Desktop\InterAutoTest\data\testdata_res.xls'
        # res = eval(pd.read_excel(r'C:\Users\Administrator\PycharmProjects\InterAutoTest\data\testdata_res.xls', sheet_name='美多商城接口测试').iloc[case_no][10])
        if os.path.exists(casepath):
            workbook = xlrd.open_workbook(casepath)
            # worksheet = workbook.sheet_by_name('美多商城接口测试')
            worksheet = workbook.sheet_by_name(sheet_name)
            ruz = worksheet.cell(case_no, 11).value
            if len(ruz)>0:
                res = eval(worksheet.cell(case_no,11).value)
                for case_id in res:
                    num = OperationExcel(casepath,sheet_name).get_row_num(case_id)
                    responsedata = eval(OperationExcel(casepath,sheet_name).get_cell_value(num,16))
                    # 找到要返回的case
                    # 找到要返回的字段
                    for ziduan in res[case_id]:
                        # print(ziduan)
                        json_expr = parse(ziduan)
                        male = json_expr.find(responsedata)
                        # male = responsedata.find(json_expr)
                        # # # 获取字段的值
                        result = [math.value for math in male][0]
                        print(result)
                        # # 利用反射赋值
                        if ziduan.find("[")!=-1:
                          a = re.search("\[(.*)\]", ziduan)
                        else:
                            a = re.search("(.*)", ziduan)
                        print(a.group(1))
                        print(type(a.group(1)))
                        # a.group(1)为字段名
                        setattr(Getdata, a.group(1), result)

    @staticmethod
    def incr14Str(string):
        """Increase one (+1) in the last number part of string."""
        rt = re.search(r'(\d+)([^\d]*$)', string)
        if rt:
            pos_left = rt.span()[0]
            num = int(rt.groups()[0])
            numStr = ''
            for i in range(len(rt.groups()[0]) - len(str(num + 1))):
                numStr += '0'
            else:
                numStr += str(num + 1)
            print(string[:pos_left] \
                  + numStr[len(numStr) - len(rt.groups()[0]):] \
                  + rt.groups()[1])
            return string[:pos_left] \
                   + numStr[len(numStr) - len(rt.groups()[0]):] \
                   + rt.groups()[1]
        else:
            raise ValueError('No suitable number segment found to +1.')

    # @staticmethod
    # def updata_inti(file_name, sheet_name, tel):
    #     # 生成随机手机号码
    #     wb = load_workbook(file_name)
    #     sheet = wb[sheet_name]
    #     sheet.cell(2, 1).value = tel
    #     wb.save(file_name)
    @staticmethod
    def updata_init(file_name, sheet_name,value):
        read_data = xlrd.open_workbook(file_name, formatting_info=True)
        write_data = copy(read_data)
        sheet_data = write_data.get_sheet(sheet_name)
        sheet_data.write(1, 1, value)
        # write_data.save(self.excel_file)
        write_data.save(r'C:\Users\wangchao\Desktop\InterAutoTest\data\testdata.xls')
        # write_data.save('../data/testdata.xls')


if __name__ == '__main__':

    s = '{"moblie":"${name}","pwd":"${pwd}"}'
    # s= '{"moblie":"556688","pwd":"995566"}'
    DoRegx().do_regx(s)
    # DoRegx().get_parmvalue()

    # s = '{"moblie":"${name}","pwd":"123456"}'
    # res = DoRegx.do_regx1(s)
    # print(res)
